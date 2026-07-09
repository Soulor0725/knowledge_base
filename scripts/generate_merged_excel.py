"""
生成合并版测试用例 Excel
将功能测试和接口测试合并到一个 Excel，按工作表区分
输出：Test_Case/智慧管理中心_测试用例.xlsx
"""
import os
import sys

# 确保能导入同目录的两个生成模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import generate_test_excel as func_mod
import generate_api_test_excel as api_mod

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Test_Case')
OUT_FILE = os.path.join(OUT_DIR, '智慧管理中心_测试用例.xlsx')

# ============ 样式（复用两个模块的样式定义）============
HEADER_FILL = func_mod.HEADER_FILL
HEADER_FONT = func_mod.HEADER_FONT
MODULE_FILL = func_mod.MODULE_FILL
MODULE_FONT = func_mod.MODULE_FONT
SUB_FILL = func_mod.SUB_FILL
SUB_FONT = func_mod.SUB_FONT
API_FILL = api_mod.API_FILL
API_FONT = api_mod.API_FONT
CELL_FONT = func_mod.CELL_FONT
MONO_FONT = api_mod.MONO_FONT
CENTER = func_mod.CENTER
LEFT = func_mod.LEFT
THIN = func_mod.THIN
BORDER = func_mod.BORDER
P0_FILL = func_mod.P0_FILL
P1_FILL = func_mod.P1_FILL
P2_FILL = func_mod.P2_FILL


def write_func_sheet(ws, cases, columns):
    """写入功能测试用例工作表"""
    ws.append(columns)
    for col in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for case in cases:
        main_module, sub_module, case_id, title, priority, case_type, pre, steps, expected = case
        row = [case_id, main_module, sub_module, title, priority, case_type, pre, steps, expected]
        ws.append(row)
        r = ws.max_row
        for col in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = LEFT if col >= 7 else CENTER
        pcell = ws.cell(row=r, column=5)
        if priority == 'P0':
            pcell.fill = P0_FILL
        elif priority == 'P1':
            pcell.fill = P1_FILL
        else:
            pcell.fill = P2_FILL

    # 合并主模块、子模块单元格
    _merge_column(ws, 2, MODULE_FILL, MODULE_FONT)
    _merge_column(ws, 3, SUB_FILL, SUB_FONT)

    # 列宽、行高
    widths = [14, 14, 12, 28, 8, 8, 22, 38, 38]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 60
    ws.freeze_panes = 'A2'


def write_api_sheet(ws, cases, columns):
    """写入接口测试用例工作表"""
    ws.append(columns)
    for col in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for case in cases:
        main_module, api, case_id, method, path, title, priority, case_type, headers, body, code, expected = case
        row = [case_id, main_module, api, method, path, title, priority, case_type, headers, body, code, expected]
        ws.append(row)
        r = ws.max_row
        for col in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = MONO_FONT if col in (5, 10) else CELL_FONT
            cell.border = BORDER
            cell.alignment = LEFT if col in (5, 9, 10, 12) else CENTER
        pcell = ws.cell(row=r, column=7)
        if priority == 'P0':
            pcell.fill = P0_FILL
        elif priority == 'P1':
            pcell.fill = P1_FILL
        else:
            pcell.fill = P2_FILL

    # 合并主模块、接口单元格
    _merge_column(ws, 2, MODULE_FILL, MODULE_FONT)
    _merge_column(ws, 3, API_FILL, API_FONT)

    widths = [18, 14, 26, 10, 40, 24, 8, 8, 24, 50, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 48
    ws.freeze_panes = 'A2'


def _merge_column(ws, col_idx, fill, font):
    """合并指定列的相同值单元格"""
    start = 2
    for i in range(3, ws.max_row + 2):
        cur = ws.cell(row=i, column=col_idx).value if i <= ws.max_row else None
        prev = ws.cell(row=i - 1, column=col_idx).value
        if i <= ws.max_row and cur == prev:
            continue
        if i - 1 > start:
            ws.merge_cells(start_row=start, start_column=col_idx, end_row=i - 1, end_column=col_idx)
            cell = ws.cell(row=start, column=col_idx)
            cell.fill = fill
            cell.font = font
            cell.alignment = CENTER
        start = i
    if ws.max_row >= start:
        ws.merge_cells(start_row=start, start_column=col_idx, end_row=ws.max_row, end_column=col_idx)
        cell = ws.cell(row=start, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER


def write_summary_sheet(ws, func_cases, api_cases):
    """写入用例统计工作表（使用公式，支持用例增删联动）"""
    from collections import defaultdict

    ws.append(['类型', '主模块', '子模块/接口数', '用例数', 'P0', 'P1', 'P2'])
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    # 先确定主模块列表（保持原 CASES 顺序去重，确保稳定）
    func_modules = []
    for c in func_cases:
        if c[0] not in func_modules:
            func_modules.append(c[0])
    api_modules = []
    for c in api_cases:
        if c[0] not in api_modules:
            api_modules.append(c[0])

    # Excel 中两个用例 sheet 的数据范围（限定行范围避免整列空值导致 MATCH 报错）
    # 功能用例：B 列主模块、C 列子模块、E 列优先级
    # 接口用例：B 列主模块、C 列接口、G 列优先级
    FUNC_RNG_MOD = "'功能测试用例'!B2:B1000"
    FUNC_RNG_SUB = "'功能测试用例'!C2:C1000"
    FUNC_RNG_PRIO = "'功能测试用例'!E2:E1000"
    API_RNG_MOD = "'接口测试用例'!B2:B1000"
    API_RNG_API = "'接口测试用例'!C2:C1000"
    API_RNG_PRIO = "'接口测试用例'!G2:G1000"

    row_idx = 2
    func_first_row = row_idx

    # 功能测试统计（公式驱动）
    for module in func_modules:
        # 子模块数：SUMPRODUCT 统计该模块下不同子模块数量（去重）
        # 原理：(模块匹配) * (子模块非空) * (该行是子模块+模块组合的首次出现)
        # 用 IFERROR 包裹避免空值导致 #N/A
        sub_formula = (
            f'=IFERROR(SUMPRODUCT(({FUNC_RNG_MOD}="{module}")*'
            f'({FUNC_RNG_SUB}<>"")*'
            f'(MATCH({FUNC_RNG_SUB}&"|"&{FUNC_RNG_MOD},{FUNC_RNG_SUB}&"|"&{FUNC_RNG_MOD},0)=ROW({FUNC_RNG_SUB})-1)),0)'
        )
        # 用例数：COUNTIF 主模块列
        total_formula = f'=COUNTIF({FUNC_RNG_MOD},"{module}")'
        # P0/P1/P2：COUNTIFS 主模块 + 优先级
        p0_formula = f'=COUNTIFS({FUNC_RNG_MOD},"{module}",{FUNC_RNG_PRIO},"P0")'
        p1_formula = f'=COUNTIFS({FUNC_RNG_MOD},"{module}",{FUNC_RNG_PRIO},"P1")'
        p2_formula = f'=COUNTIFS({FUNC_RNG_MOD},"{module}",{FUNC_RNG_PRIO},"P2")'

        ws.append(['功能测试', module, sub_formula, total_formula, p0_formula, p1_formula, p2_formula])
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = CENTER
        row_idx += 1

    func_last_row = row_idx - 1
    # 功能测试小计（SUM 公式）
    ws.append([
        '功能测试小计', '—',
        f'=SUM(C{func_first_row}:C{func_last_row})',
        f'=SUM(D{func_first_row}:D{func_last_row})',
        f'=SUM(E{func_first_row}:E{func_last_row})',
        f'=SUM(F{func_first_row}:F{func_last_row})',
        f'=SUM(G{func_first_row}:G{func_last_row})',
    ])
    for col in range(1, 8):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = SUB_FILL
        cell.font = SUB_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    func_subtotal_row = row_idx
    row_idx += 1

    # 接口测试统计（公式驱动）
    api_first_row = row_idx
    for module in api_modules:
        # 子接口数：该模块下不同接口数量（去重）
        api_formula = (
            f'=IFERROR(SUMPRODUCT(({API_RNG_MOD}="{module}")*'
            f'({API_RNG_API}<>"")*'
            f'(MATCH({API_RNG_API}&"|"&{API_RNG_MOD},{API_RNG_API}&"|"&{API_RNG_MOD},0)=ROW({API_RNG_API})-1)),0)'
        )
        total_formula = f'=COUNTIF({API_RNG_MOD},"{module}")'
        p0_formula = f'=COUNTIFS({API_RNG_MOD},"{module}",{API_RNG_PRIO},"P0")'
        p1_formula = f'=COUNTIFS({API_RNG_MOD},"{module}",{API_RNG_PRIO},"P1")'
        p2_formula = f'=COUNTIFS({API_RNG_MOD},"{module}",{API_RNG_PRIO},"P2")'

        ws.append(['接口测试', module, api_formula, total_formula, p0_formula, p1_formula, p2_formula])
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = CENTER
        row_idx += 1

    api_last_row = row_idx - 1
    # 接口测试小计
    ws.append([
        '接口测试小计', '—',
        f'=SUM(C{api_first_row}:C{api_last_row})',
        f'=SUM(D{api_first_row}:D{api_last_row})',
        f'=SUM(E{api_first_row}:E{api_last_row})',
        f'=SUM(F{api_first_row}:F{api_last_row})',
        f'=SUM(G{api_first_row}:G{api_last_row})',
    ])
    for col in range(1, 8):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = SUB_FILL
        cell.font = SUB_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    api_subtotal_row = row_idx
    row_idx += 1

    # 总计
    ws.append([
        '总计', '—', '—',
        f'=D{func_subtotal_row}+D{api_subtotal_row}',
        f'=E{func_subtotal_row}+E{api_subtotal_row}',
        f'=F{func_subtotal_row}+F{api_subtotal_row}',
        f'=G{func_subtotal_row}+G{api_subtotal_row}',
    ])
    for col in range(1, 8):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = MODULE_FILL
        cell.font = MODULE_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    # 合并"类型"列：功能测试多行合并、接口测试多行合并
    if func_last_row > func_first_row:
        ws.merge_cells(start_row=func_first_row, start_column=1, end_row=func_last_row, end_column=1)
        cell = ws.cell(row=func_first_row, column=1)
        cell.fill = MODULE_FILL
        cell.font = MODULE_FONT
        cell.alignment = CENTER
    if api_last_row > api_first_row:
        ws.merge_cells(start_row=api_first_row, start_column=1, end_row=api_last_row, end_column=1)
        cell = ws.cell(row=api_first_row, column=1)
        cell.fill = MODULE_FILL
        cell.font = MODULE_FONT
        cell.alignment = CENTER

    widths = [16, 16, 14, 12, 8, 8, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[1].height = 30


def build_merged_excel():
    wb = Workbook()

    # Sheet1: 功能测试用例
    ws_func = wb.active
    ws_func.title = '功能测试用例'
    write_func_sheet(ws_func, func_mod.CASES, func_mod.COLUMNS)

    # Sheet2: 接口测试用例
    ws_api = wb.create_sheet('接口测试用例')
    write_api_sheet(ws_api, api_mod.CASES, api_mod.COLUMNS)

    # Sheet3: 用例统计
    ws_summary = wb.create_sheet('用例统计')
    write_summary_sheet(ws_summary, func_mod.CASES, api_mod.CASES)

    wb.save(OUT_FILE)
    print(f'已生成: {OUT_FILE}')
    print(f'功能测试用例: {len(func_mod.CASES)}')
    print(f'接口测试用例: {len(api_mod.CASES)}')
    print(f'总计: {len(func_mod.CASES) + len(api_mod.CASES)}')
    print(f'工作表: 功能测试用例 / 接口测试用例 / 用例统计')


if __name__ == '__main__':
    build_merged_excel()
